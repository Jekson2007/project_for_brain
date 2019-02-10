#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import run


file_for_analiz = input('Enter name file and that adress directory exel file  :  ')
open_file = load_workbook(filename= file_for_analiz)
give_name_list = open_file.sheetnames[0] # call(name) first sheet our exelbook
values = open_file[give_name_list]
number_of_elements_in_the_sample = int(values['C2'].value) # number_of_elements_in_the_sample
n = number_of_elements_in_the_sample + 2 # correct number_of_elements_in_the_sample
first_sample = [] # make list with first sample
for value in range(2,n): 
    cell_name = 'A{}'.format(value) # cell name in excel file
    first_sample.append(values[cell_name].value) # add next element in sample in excel file

second_sample = [] # make list with second sample
for value in range(2,n):
    cell_name = 'B{}'.format(value) # cell name in excel file
    second_sample.append(values[cell_name].value) # add next element in sample in excel file

run.drawGrafic(first_sample,second_sample) # run 'run' file 
